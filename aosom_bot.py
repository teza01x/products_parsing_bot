import time
import warnings
import pandas as pd
import telebot
import sqlite3
from selenium.webdriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from config import *


class Objects:
    def __init__(self, wait):
        self.wait = wait


    def check_for_general_info(self, general_selector):
        try:
            self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, general_selector)))
            return True
        except:
            return False


    def instock_status(self, buy_btn):
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, buy_btn)))
            return True
        except:
            return False


    def parse_info_from_page(self, browser, right_menu_class, title_class, price_class, alt_price_class):
        html = browser.page_source
        soup = BeautifulSoup(html, 'html.parser')
        right_menu = soup.find('div', class_=right_menu_class)
        try:
            title = right_menu.find('h1', class_=title_class).get_text()
            title = " ".join([i for i in [i.replace('\n', '') for i in [i for i in title.split(" ") if len(i) > 0]] if len(i) > 0])
        except:
            title = 'None'
        try:
            try:
                price = right_menu.find('div', class_=price_class).get_text()
            except:
                price = right_menu.find('div', class_=alt_price_class).get_text()
            price = price.replace('CA$', '')
            price = price.replace(',', '')
            price = float(price)
        except:
            price = 'None'
        try:
            td_element = right_menu.find('td', text='SKU')
            sku = td_element.find_next('td').text.strip()
        except:
            sku = 'None'
        return title, price, sku


def create_new_excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(['SKU', 'PRICE', 'STOCK STATUS', 'TITLE', 'LINK'])

    now = datetime.now()
    current_excel_name = now.strftime("%d_%m_%y_%H.%M")
    workbook.save("out/{}.xlsx".format(current_excel_name))
    return "out/{}.xlsx".format(current_excel_name)


def add_new_data(data, excel_name):
    workbook = load_workbook(excel_name)

    sheet = workbook.active

    for i in data:
        sheet.append(i)

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error {}".format(e))


def add_data_with_stock_and_price_change(data, excel_name, color_price, color_stock):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill1 = PatternFill(start_color=color_price, end_color=color_price, fill_type='solid')
    fill2 = PatternFill(start_color=color_stock, end_color=color_stock, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color1 = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color1.fill = fill1

        cell_to_color2 = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color2.fill = fill2

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))


def add_data_with_stock_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=3)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))



def add_data_with_price_change(data, excel_name, current_color):
    workbook = load_workbook(excel_name)
    sheet = workbook.active

    fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')

    for row_data in data:
        sheet.append(row_data)
        cell_to_color = sheet.cell(row=sheet.max_row, column=2)
        cell_to_color.fill = fill

    try:
        workbook.save(excel_name)
    except Exception as e:
        print("Error: {}".format(e))


def item_exists(sku):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    result = cursor.execute("SELECT sku FROM item_info WHERE sku = ?", (sku,))
    exists = bool(len(result.fetchall()))

    conn.close()

    return exists


def add_item_to_db(sku, price, stock):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    cursor.execute("INSERT INTO item_info (sku, price, stock) VALUES(?, ?, ?)", (sku, price, stock,))

    conn.commit()
    conn.close()


def get_item_info(sku):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    result = cursor.execute("SELECT price, stock FROM item_info WHERE sku = ?", (sku,))
    result = result.fetchall()[0]
    conn.close()

    return result


def update_item_info(sku, stock_status, item_price):
    conn = sqlite3.connect('items.db')
    cursor = conn.cursor()

    cursor.execute("UPDATE item_info SET price = ?, stock = ? WHERE sku = ?", (item_price, stock_status, sku,))

    conn.commit()
    conn.close()



def telegram_msg(bot_status):
    bot = telebot.TeleBot(telegram_token)
    if bot_status == "start":
        bot.send_message(telegram_user_id, "🟢 The bot started scraping pages from website. 🟢\nDo not open the excel spreadsheet that the bot is currently working with, otherwise some information may be lost.")
    elif bot_status == "end":
        bot.send_message(telegram_user_id, '♦️ The bot has finished parsing pages from website. ♦️\nThe bot has stopped working️, now you can work with excel spreadsheets. (folder "out")')


def main(browser, link, current_excel_name):
    wait = WebDriverWait(browser, 10)
    html_elements = {
        'general_info_css': '#js-product-content > div.product-right',
        'buy_btn': '//button[@class="product-right-operate-buy a-btn a-btn--primary"]',
        'right_menu_class': 'product-right',
        'title_class': 'title-name a-fonts--26 a-colors--black-t js-detail-title',
        'price_class': 'price-now a-fonts--w-700 red',
        'alt_price_class': 'price-now a-fonts--w-700',
        'sku_code': '#js-product-content > div.product-right > div:nth-child(19) > div.product-right-item-wrap > div.product-right-item-table-wrap.a-fonts--16.a-fonts--w-500.mt-4 > table > tbody > tr:nth-child(2) > td:nth-child(2)',
    }
    elem = html_elements
    objects = Objects(wait)
    try:
        if objects.check_for_general_info(elem['general_info_css']) == True:
            item_info_list = list()
            item_title, item_price, item_sku = objects.parse_info_from_page(browser, elem['right_menu_class'], elem['title_class'], elem['price_class'], elem['alt_price_class'])
            if objects.instock_status(elem['buy_btn']) == True:
                item_stock = 'IN STOCK'
            else:
                item_stock = 'OUT OF STOCK'
            item_info_list.append(item_sku)
            item_info_list.append(item_price)
            item_info_list.append(item_stock)
            item_info_list.append(item_title)
            item_info_list.append(link)

            if item_exists(item_sku) == False:
                add_item_to_db(item_sku, item_price, item_stock)
                add_new_data([item_info_list], current_excel_name)
            elif item_exists(item_sku) == True:
                prev_item_price, prev_item_stock = get_item_info(item_sku)
                if prev_item_price == item_price and prev_item_stock == item_stock:
                    add_new_data([item_info_list], current_excel_name)
                elif prev_item_stock != item_stock and prev_item_price == item_price:
                    add_data_with_stock_change([item_info_list], current_excel_name, red_color)
                    update_item_info(item_sku, item_stock, item_price)
                elif prev_item_stock == item_stock and prev_item_price != item_price:
                    add_data_with_price_change([item_info_list], current_excel_name, yellow_color)
                    update_item_info(item_sku, item_stock, item_price)
                elif prev_item_stock != item_stock and prev_item_price != item_price:
                    add_data_with_stock_and_price_change([item_info_list], current_excel_name, yellow_color, red_color)
                    update_item_info(item_sku, item_stock, item_price)
            time.sleep(10)
    except:
        time.sleep(10)


if __name__ == "__main__":
    options = Options()
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    options.binary_location = ""
    try:
        browser = Chrome("chromedriver.exe", chrome_options=options)
        browser.maximize_window()

        current_excel_name = create_new_excel()

        df = pd.read_excel(aosom_link_excel_file)
        link_column = df['Supplier Product Link']

        telegram_msg("start")

        for i in range(len(link_column)):
            link = link_column.iloc[i]
            browser.get(link)
            main(browser, link, current_excel_name)

        telegram_msg("end")
        browser.quit()
    except:
        telegram_msg("end")